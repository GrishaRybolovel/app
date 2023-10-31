import asyncio
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from typing import List, Any, Dict
from datetime import datetime

import os
print(os.getcwd())
print(os.listdir())

app = FastAPI()

origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Load the Excel workbook
wb = load_workbook('data.xlsx', data_only=True)
ws = wb.active

# A utility function to fetch all data from the Excel file
def get_excel_data_sync(start_date, end_date):
    wb_sales = wb['Продажи']
    wb_loss = wb['Потери']
    wb_abil = wb['Навыки']

    print(start_date, end_date)

    data = {}
    for row in wb_sales.iter_rows(values_only=True):
        if isinstance(row[2], int) and datetime.strptime(start_date, '%Y-%m-%d') <= row[0] <= datetime.strptime(end_date, '%Y-%m-%d'):
            data[row[0]] = {}
            data[row[0]][row[1]] = [row[2], '', '']

    for row in wb_loss.iter_rows(values_only=True):
        if isinstance(row[2], int) and datetime.strptime(start_date, '%Y-%m-%d') <= row[0] <= datetime.strptime(end_date, '%Y-%m-%d'):
            data[row[0]][row[1]][1] = row[2]

    for row in wb_abil.iter_rows(values_only=True):
        if isinstance(row[2], int) and datetime.strptime(start_date, '%Y-%m-%d') <= row[0] <= datetime.strptime(end_date, '%Y-%m-%d'):
            data[row[0]][row[1]][2] = row[2]

    return data

# Asynchronous function that runs our synchronous Excel data fetcher
async def get_excel_data(start_date, end_date):
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, get_excel_data_sync, start_date, end_date)

@app.get("/data/")
async def read_data(date_start = '2023-01-01', date_end = '2023-01-05'):
    date_start = date_start[:10]
    date_end = date_end[:10]
    data = await get_excel_data(date_start, date_end)
    if not data:
        raise HTTPException(status_code=404, detail="Data not found")
    return data